"""
Script para crear archivos Excel y CSV vacíos con estructura definida
"""

import sys
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Instalando openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

# Definir columnas para cada archivo
VENTAS_VIRTUALES_COLUMNS = [
    "id",
    "categoria",
    "fecha",
    "canal_venta_virtual",
    "medio_pago",
    "pago_tipo",
    "precio_unitario",
    "precio_promocional_preventa_40off",
    "precio_venta",
    "cantidad",
    "total",
    "vendedor_o_referido",
    "comprador_email",
    "comprador_whatsapp",
    "estado",
    "notas"
]

VENTAS_FISICAS_COLUMNS = [
    "id",
    "categoria",
    "fecha",
    "canal_venta_p2p",
    "medio_pago",
    "precio_unitario",
    "precio_promocional_preventa_40off",
    "precio_venta",
    "cantidad",
    "total",
    "vendedor_o_referido",
    "comprador_email",
    "comprador_whatsapp",
    "asiento_fisico_numero",
    "asiento_estado",
    "estado_pago",
    "notas"
]

RESERVAS_ASIENTOS_COLUMNS = [
    "id",
    "categoria",
    "fecha",
    "vendedor_o_referido",
    "comprador_email",
    "comprador_whatsapp",
    "asiento_fisico_numero",
    "asiento_estado",
    "monto_reserva",
    "medio_pago",
    "estado_pago",
    "observaciones"
]


def create_excel_file(filename: str, columns: list, categoria: str):
    """Crear archivo Excel con hoja 'datos' y encabezados"""
    wb = Workbook()
    ws = wb.active
    ws.title = "datos"
    
    # Estilo para encabezados
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Escribir encabezados
    for col_idx, header in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Ajustar ancho de columnas
    for col_idx in range(1, len(columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 25
    
    # Guardar
    filepath = Path("data") / filename
    wb.save(filepath)
    print(f"Creado: {filepath}")


def create_csv_file(filename: str, columns: list):
    """Crear archivo CSV con encabezados"""
    filepath = Path("data") / filename
    with open(filepath, 'w', encoding='utf-8', newline='') as f:
        f.write(','.join(columns) + '\n')
    print(f"Creado: {filepath}")


def main():
    """Crear todos los archivos"""
    print("Creando archivos de datos...")
    print("=" * 50)
    
    # Ventas Virtuales
    create_excel_file("ventas_virtuales.xlsx", VENTAS_VIRTUALES_COLUMNS, "virtual")
    create_csv_file("ventas_virtuales.csv", VENTAS_VIRTUALES_COLUMNS)
    
    # Ventas Físicas
    create_excel_file("ventas_fisicas.xlsx", VENTAS_FISICAS_COLUMNS, "fisica")
    create_csv_file("ventas_fisicas.csv", VENTAS_FISICAS_COLUMNS)
    
    # Reservas Asientos
    create_excel_file("reservas_asientos.xlsx", RESERVAS_ASIENTOS_COLUMNS, "reserva")
    create_csv_file("reservas_asientos.csv", RESERVAS_ASIENTOS_COLUMNS)
    
    print("=" * 50)
    print("Todos los archivos creados exitosamente")


if __name__ == "__main__":
    main()

