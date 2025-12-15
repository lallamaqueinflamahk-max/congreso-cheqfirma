"""
Generador de reporte Excel diario
6 hojas: RESUMEN, VENDIDOS, RESERVADOS, VACANTES, CANCELACIONES, AUDIT
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict, Any
from datetime import datetime, timedelta
from report.policy import Seat, SeatStatus, VentaOrigen
from report.metrics import MetricsCalculator


class ExcelReportGenerator:
    """Generador de reportes Excel"""
    
    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remover hoja por defecto
    
    def generate(
        self, 
        seats: List[Seat], 
        metrics: Dict[str, Any],
        audit_log: List[Dict[str, Any]],
        output_path: str
    ):
        """
        Generar reporte Excel completo
        
        Args:
            seats: Lista de todos los asientos
            metrics: Métricas calculadas
            audit_log: Log de auditoría
            output_path: Ruta donde guardar el Excel
        """
        # Hoja 1: RESUMEN GENERAL
        self._create_resumen_sheet(metrics)
        
        # Hoja 2: VENDIDOS
        self._create_vendidos_sheet(seats)
        
        # Hoja 3: RESERVADOS
        self._create_reservados_sheet(seats)
        
        # Hoja 4: VACANTES
        self._create_vacantes_sheet(seats)
        
        # Hoja 5: CANCELACIONES
        self._create_cancelaciones_sheet(seats)
        
        # Hoja 6: AUDIT
        self._create_audit_sheet(audit_log)
        
        # Guardar
        self.wb.save(output_path)
    
    def _create_resumen_sheet(self, metrics: Dict[str, Any]):
        """Hoja 1: RESUMEN GENERAL"""
        ws = self.wb.create_sheet("RESUMEN GENERAL", 0)
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "REPORTE DIARIO - CONGRESO ADN HUMANO"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        # Capacidad y Ocupación
        ws[f'A{row}'] = "CAPACIDAD TOTAL"
        ws[f'B{row}'] = metrics["total"]["capacidad_total"]
        row += 1
        
        ws[f'A{row}'] = "VENDIDOS"
        ws[f'B{row}'] = metrics["total"]["vendidos"]
        ws[f'C{row}'] = f"Total: {metrics['total']['vendidos']} | WhatsApp: {metrics['whatsapp_caliente']['vendidos']} | Virtual: {metrics['virtual_frio']['vendidos']}"
        row += 1
        
        ws[f'A{row}'] = "RESERVADOS"
        ws[f'B{row}'] = metrics["total"]["reservados"]
        ws[f'C{row}'] = f"Total: {metrics['total']['reservados']} | WhatsApp: {metrics['whatsapp_caliente']['reservados']} | Virtual: {metrics['virtual_frio']['reservados']}"
        row += 1
        
        ws[f'A{row}'] = "VACANTES"
        ws[f'B{row}'] = metrics["total"]["vacantes"]
        row += 2
        
        # Ingresos
        ws[f'A{row}'] = "INGRESOS COBRADOS"
        ws[f'A{row}'].font = title_font
        row += 1
        
        ws[f'A{row}'] = "Total"
        ws[f'B{row}'] = metrics["total"]["ingresos_cobrados"]
        row += 1
        
        ws[f'A{row}'] = "WhatsApp Caliente"
        ws[f'B{row}'] = metrics["whatsapp_caliente"]["ingresos_cobrados"]
        row += 1
        
        ws[f'A{row}'] = "Virtual Frío"
        ws[f'B{row}'] = metrics["virtual_frio"]["ingresos_cobrados"]
        row += 2
        
        # Señado y Saldos
        ws[f'A{row}'] = "SEÑADO TOTAL"
        ws[f'B{row}'] = metrics["total"]["senado_total"]
        row += 1
        
        ws[f'A{row}'] = "SALDO PENDIENTE"
        ws[f'B{row}'] = metrics["total"]["saldo_pendiente"]
        row += 1
        
        ws[f'A{row}'] = "CANCELACIONES"
        ws[f'B{row}'] = metrics["total"]["cancelaciones"]
        row += 2
        
        # Alertas
        if metrics["alertas"]:
            ws[f'A{row}'] = "ALERTAS AUTOMÁTICAS"
            ws[f'A{row}'].font = title_font
            row += 1
            
            for alert in metrics["alertas"]:
                ws[f'A{row}'] = f"{alert['tipo']} ({alert['severidad']})"
                ws[f'B{row}'] = f"Cantidad: {alert.get('cantidad', 'N/A')}"
                if 'asientos' in alert:
                    ws[f'C{row}'] = f"Asientos: {', '.join(map(str, alert['asientos']))}"
                row += 1
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50
    
    def _create_vendidos_sheet(self, seats: List[Seat]):
        """Hoja 2: VENDIDOS"""
        ws = self.wb.create_sheet("VENDIDOS")
        
        vendidos = [s for s in seats if s.status == SeatStatus.VENDIDO]
        
        # Headers
        headers = [
            "Asiento", "Comprador", "Email", "Teléfono",
            "Monto", "Fecha Pago", "Origen de Venta", "Tipo"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, seat in enumerate(vendidos, 2):
            ws.cell(row=row_idx, column=1, value=seat.seat_number)
            ws.cell(row=row_idx, column=2, value=seat.buyer_name or "")
            ws.cell(row=row_idx, column=3, value=seat.buyer_email or "")
            ws.cell(row=row_idx, column=4, value=seat.buyer_phone or "")
            ws.cell(row=row_idx, column=5, value=seat.amount_paid)
            ws.cell(row=row_idx, column=6, value=seat.paid_at.strftime("%Y-%m-%d %H:%M") if seat.paid_at else "")
            ws.cell(row=row_idx, column=7, value=seat.venta_origen.value if seat.venta_origen else "")
            ws.cell(row=row_idx, column=8, value="WhatsApp" if seat.venta_origen == VentaOrigen.WHATSAPP_CALIENTE else "Virtual")
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_reservados_sheet(self, seats: List[Seat]):
        """Hoja 3: RESERVADOS"""
        ws = self.wb.create_sheet("RESERVADOS")
        
        reservados = [s for s in seats if s.status == SeatStatus.RESERVADO]
        
        # Headers
        headers = [
            "Asiento", "Comprador", "Email", "Origen de Venta",
            "Fecha Reserva", "Vencimiento", "Seña Abonada",
            "Saldo Pendiente", "Riesgo Liberación"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, seat in enumerate(reservados, 2):
            saldo_pendiente = seat.total_amount - seat.amount_paid
            tiene_sena = seat.deposit_amount > 0
            riesgo = "SÍ" if (seat.expires_at and seat.expires_at < datetime.now() + timedelta(days=2)) else "NO"
            
            ws.cell(row=row_idx, column=1, value=seat.seat_number)
            ws.cell(row=row_idx, column=2, value=seat.buyer_name or "")
            ws.cell(row=row_idx, column=3, value=seat.buyer_email or "")
            ws.cell(row=row_idx, column=4, value=seat.venta_origen.value if seat.venta_origen else "")
            ws.cell(row=row_idx, column=5, value=seat.reserved_at.strftime("%Y-%m-%d %H:%M") if seat.reserved_at else "")
            ws.cell(row=row_idx, column=6, value=seat.expires_at.strftime("%Y-%m-%d %H:%M") if seat.expires_at else "")
            ws.cell(row=row_idx, column=7, value="SÍ" if tiene_sena else "NO")
            ws.cell(row=row_idx, column=8, value=saldo_pendiente)
            ws.cell(row=row_idx, column=9, value=riesgo)
            
            # Resaltar riesgo
            if riesgo == "SÍ":
                ws.cell(row=row_idx, column=9).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
                ws.cell(row=row_idx, column=9).font = Font(bold=True, color="FFFFFF")
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_vacantes_sheet(self, seats: List[Seat]):
        """Hoja 4: VACANTES"""
        ws = self.wb.create_sheet("VACANTES")
        
        vacantes = [s for s in seats if s.status == SeatStatus.VACANTE]
        
        # Headers
        headers = ["Asiento", "Sección", "Fila", "Número"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, seat in enumerate(vacantes, 2):
            ws.cell(row=row_idx, column=1, value=seat.seat_number)
            ws.cell(row=row_idx, column=2, value=seat.section or "")
            ws.cell(row=row_idx, column=3, value=seat.row or "")
            ws.cell(row=row_idx, column=4, value=seat.seat_number)
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    def _create_cancelaciones_sheet(self, seats: List[Seat]):
        """Hoja 5: CANCELACIONES"""
        ws = self.wb.create_sheet("CANCELACIONES")
        
        cancelados = [s for s in seats if s.cancellation_reason]
        
        # Headers
        headers = [
            "Asiento", "Origen de Venta", "Motivo",
            "Monto Abonado", "Reembolso Posible", "Reembolsado"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, seat in enumerate(cancelados, 2):
            reembolso_posible = "SÍ" if seat.refund_amount > 0 else "NO"
            reembolsado = "SÍ" if seat.refunded else "NO"
            
            ws.cell(row=row_idx, column=1, value=seat.seat_number)
            ws.cell(row=row_idx, column=2, value=seat.venta_origen.value if seat.venta_origen else "")
            ws.cell(row=row_idx, column=3, value=seat.cancellation_reason or "")
            ws.cell(row=row_idx, column=4, value=seat.amount_paid)
            ws.cell(row=row_idx, column=5, value=reembolso_posible)
            ws.cell(row=row_idx, column=6, value=reembolsado)
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_audit_sheet(self, audit_log: List[Dict[str, Any]]):
        """Hoja 6: AUDIT"""
        ws = self.wb.create_sheet("AUDIT")
        
        # Headers
        headers = [
            "Fecha", "Acción Automática", "Asiento",
            "Motivo", "Estado Anterior", "Estado Nuevo"
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="A5A5A5", end_color="A5A5A5", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
        
        # Datos
        for row_idx, entry in enumerate(audit_log, 2):
            ws.cell(row=row_idx, column=1, value=entry["fecha"].strftime("%Y-%m-%d %H:%M:%S") if isinstance(entry["fecha"], datetime) else str(entry["fecha"]))
            ws.cell(row=row_idx, column=2, value=entry["accion"])
            ws.cell(row=row_idx, column=3, value=entry["asiento"])
            ws.cell(row=row_idx, column=4, value=entry["motivo"])
            ws.cell(row=row_idx, column=5, value=entry["estado_anterior"])
            ws.cell(row=row_idx, column=6, value=entry["estado_nuevo"])
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20

