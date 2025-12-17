"""
Sistema automático de reporte de ventas "desde el inicio hasta hoy"
Genera Excel completo con datos actualizados de ventas, reservas e incidencias
Incluye estadísticas, probabilidades simples y gráficos
"""

import sys
import os
import json
import yaml
from pathlib import Path
from datetime import datetime, timedelta
from typing import Dict, List, Any, Optional
from collections import defaultdict
import logging

# Agregar el directorio report al path
sys.path.insert(0, str(Path(__file__).parent.parent))

from report.policy import Seat, SeatStatus, VentaOrigen
from report.datasources.factory import create_data_adapter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.axis import DateAxis
from openpyxl.utils import get_column_letter
import numpy as np
from scipy import stats

# Configurar logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class SalesReportGenerator:
    """Generador de reporte completo de ventas"""
    
    def __init__(self, config: Dict[str, Any]):
        self.config = config
        self.wb = Workbook()
        self.wb.remove(self.wb.active)  # Remover hoja por defecto
        self.seats: List[Seat] = []
        self.metrics: Dict[str, Any] = {}
        
    def generate(self) -> str:
        """Generar reporte completo y retornar ruta del archivo"""
        logger.info("Iniciando generación de reporte de ventas...")
        
        # 1. Cargar datos
        self._load_data()
        
        # 2. Calcular métricas
        self._calculate_metrics()
        
        # 3. Crear pestañas
        self._create_resumen_sheet()
        self._create_ventas_detalle_sheet()
        self._create_reservas_detalle_sheet()
        self._create_problemas_detalle_sheet()
        self._create_proyeccion_sheet()
        
        # 4. Guardar archivo
        output_dir = Path(self.config.get("output_dir", "report/output"))
        output_dir.mkdir(parents=True, exist_ok=True)
        
        fecha = datetime.now()
        filename = f"REPORTE_VENTAS_{fecha.strftime('%Y-%m-%d')}.xlsx"
        filepath = output_dir / filename
        
        self.wb.save(filepath)
        logger.info(f"Reporte guardado en: {filepath}")
        
        # 5. Guardar metadata
        self._save_metadata(output_dir, fecha)
        
        return str(filepath)
    
    def _load_data(self):
        """Cargar datos desde fuente configurada"""
        logger.info("Cargando datos...")
        adapter = create_data_adapter(self.config["datasource"])
        self.seats = adapter.get_all_seats()
        logger.info(f"Cargados {len(self.seats)} asientos")
    
    def _calculate_metrics(self):
        """Calcular todas las métricas y KPIs"""
        logger.info("Calculando métricas...")
        
        # Separar ventas físicas y virtuales
        ventas_fisicas = [
            s for s in self.seats 
            if s.status == SeatStatus.VENDIDO 
            and (s.venta_origen == VentaOrigen.WHATSAPP_CALIENTE or s.venta_origen is None)
        ]
        
        ventas_virtuales = [
            s for s in self.seats 
            if s.status == SeatStatus.VENDIDO 
            and s.venta_origen == VentaOrigen.VIRTUAL_FRIO
        ]
        
        reservas = [s for s in self.seats if s.status == SeatStatus.RESERVADO]
        reservas_confirmadas = [s for s in reservas if s.amount_paid >= s.total_amount * 0.5]
        
        # KPIs
        ventas_total = len(ventas_fisicas) + len(ventas_virtuales)
        ingresos_total = sum(s.amount_paid for s in ventas_fisicas + ventas_virtuales + reservas)
        ticket_promedio = ingresos_total / ventas_total if ventas_total > 0 else 0
        
        # Tasa de conversión reservas -> venta
        total_reservas_hist = len(reservas) + ventas_total
        tasa_conversion = (ventas_total / total_reservas_hist * 100) if total_reservas_hist > 0 else 0
        
        # Top productos/eventos (agrupar por tipo)
        productos_count = defaultdict(int)
        for s in ventas_fisicas + ventas_virtuales:
            tipo = "Físico" if s in ventas_fisicas else "Virtual"
            productos_count[tipo] += 1
        
        self.metrics = {
            "ventas_total": ventas_total,
            "ventas_fisico": len(ventas_fisicas),
            "ventas_virtual": len(ventas_virtuales),
            "reservas_total": len(reservas),
            "reservas_confirmadas": len(reservas_confirmadas),
            "tasa_conversion": tasa_conversion,
            "ingresos": ingresos_total,
            "ticket_promedio": ticket_promedio,
            "top_productos": dict(sorted(productos_count.items(), key=lambda x: x[1], reverse=True)),
            "ventas_por_dia": self._calculate_sales_by_day(),
            "ventas_fisico_por_dia": self._calculate_sales_by_day(fisico=True),
            "ventas_virtual_por_dia": self._calculate_sales_by_day(fisico=False),
        }
    
    def _calculate_sales_by_day(self, fisico: Optional[bool] = None) -> Dict[str, int]:
        """Calcular ventas por día"""
        sales_by_day = defaultdict(int)
        
        for seat in self.seats:
            if seat.status == SeatStatus.VENDIDO and seat.paid_at:
                # Filtrar por tipo si se especifica
                if fisico is not None:
                    is_fisico = seat.venta_origen != VentaOrigen.VIRTUAL_FRIO
                    if fisico != is_fisico:
                        continue
                
                fecha_str = seat.paid_at.strftime("%Y-%m-%d")
                sales_by_day[fecha_str] += 1
        
        return dict(sorted(sales_by_day.items()))
    
    def _create_resumen_sheet(self):
        """Pestaña Resumen: KPIs y gráficos"""
        ws = self.wb.create_sheet("Resumen", 0)
        
        # Estilos
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14)
        kpi_font = Font(bold=True, size=11)
        
        row = 1
        
        # Título
        ws.merge_cells(f'A{row}:C{row}')
        ws[f'A{row}'] = "REPORTE DE VENTAS - CONGRESO ADN HUMANO"
        ws[f'A{row}'].font = title_font
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        row += 2
        
        # KPIs
        ws[f'A{row}'] = "KPIs PRINCIPALES"
        ws[f'A{row}'].font = title_font
        row += 1
        
        kpis = [
            ("Ventas Total", self.metrics["ventas_total"]),
            ("Ventas Físico", self.metrics["ventas_fisico"]),
            ("Ventas Virtual", self.metrics["ventas_virtual"]),
            ("Reservas Totales", self.metrics["reservas_total"]),
            ("Reservas Confirmadas", self.metrics["reservas_confirmadas"]),
            ("Tasa Conversión Reservas->Venta", f"{self.metrics['tasa_conversion']:.1f}%"),
            ("Ingresos Total", f"Gs {self.metrics['ingresos']:,.0f}"),
            ("Ticket Promedio", f"Gs {self.metrics['ticket_promedio']:,.0f}"),
        ]
        
        for label, value in kpis:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = kpi_font
            ws[f'B{row}'] = value
            row += 1
        
        row += 1
        
        # Top Productos/Eventos
        ws[f'A{row}'] = "TOP PRODUCTOS/EVENTOS"
        ws[f'A{row}'].font = title_font
        row += 1
        
        for producto, cantidad in list(self.metrics["top_productos"].items())[:5]:
            ws[f'A{row}'] = producto
            ws[f'B{row}'] = cantidad
            row += 1
        
        row += 2
        
        # Datos para gráficos
        chart_data_row = row
        ws[f'A{row}'] = "Fecha"
        ws[f'B{row}'] = "Ventas por Día"
        ws[f'C{row}'] = "Físico"
        ws[f'D{row}'] = "Virtual"
        ws[f'E{row}'] = "Acumulado"
        
        for cell in [f'A{row}', f'B{row}', f'C{row}', f'D{row}', f'E{row}']:
            ws[cell].font = header_font
            ws[cell].fill = header_fill
        
        row += 1
        
        # Llenar datos de ventas por día
        acumulado = 0
        for fecha, cantidad in self.metrics["ventas_por_dia"].items():
            acumulado += cantidad
            fisico = self.metrics["ventas_fisico_por_dia"].get(fecha, 0)
            virtual = self.metrics["ventas_virtual_por_dia"].get(fecha, 0)
            
            ws[f'A{row}'] = fecha
            ws[f'B{row}'] = cantidad
            ws[f'C{row}'] = fisico
            ws[f'D{row}'] = virtual
            ws[f'E{row}'] = acumulado
            row += 1
        
        # Gráfico 1: Línea de ventas por día
        chart1 = LineChart()
        chart1.title = "Ventas por Día"
        chart1.style = 10
        chart1.y_axis.title = "Cantidad"
        chart1.x_axis.title = "Fecha"
        
        data = Reference(ws, min_col=2, min_row=chart_data_row, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=row-1)
        chart1.add_data(data, titles_from_data=True)
        chart1.set_categories(cats)
        ws.add_chart(chart1, f"G{chart_data_row}")
        
        # Gráfico 2: Barras Físico vs Virtual
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 10
        chart2.title = "Ventas Físico vs Virtual"
        chart2.y_axis.title = "Cantidad"
        chart2.x_axis.title = "Fecha"
        
        data = Reference(ws, min_col=3, min_row=chart_data_row, max_col=4, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=row-1)
        chart2.add_data(data, titles_from_data=True)
        chart2.set_categories(cats)
        ws.add_chart(chart2, f"G{chart_data_row + 15}")
        
        # Gráfico 3: Acumulado
        chart3 = LineChart()
        chart3.title = "Ventas Acumuladas"
        chart3.style = 10
        chart3.y_axis.title = "Cantidad Acumulada"
        chart3.x_axis.title = "Fecha"
        
        data = Reference(ws, min_col=5, min_row=chart_data_row, max_row=row-1)
        cats = Reference(ws, min_col=1, min_row=chart_data_row+1, max_row=row-1)
        chart3.add_data(data, titles_from_data=True)
        chart3.set_categories(cats)
        ws.add_chart(chart3, f"G{chart_data_row + 30}")
        
        # Ajustar columnas
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
    
    def _create_ventas_detalle_sheet(self):
        """Pestaña Ventas_detalle"""
        ws = self.wb.create_sheet("Ventas_detalle")
        
        # Headers
        headers = [
            "Fecha", "Canal", "Producto/Evento", "Cantidad", 
            "Precio", "Total", "Método de Pago", "Estado", 
            "Referencia/ID", "Vendedor"
        ]
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        row = 2
        ventas = [s for s in self.seats if s.status == SeatStatus.VENDIDO]
        
        for seat in ventas:
            canal = "Físico" if seat.venta_origen != VentaOrigen.VIRTUAL_FRIO else "Virtual"
            producto = "Congreso Presencial" if canal == "Físico" else "Congreso Virtual"
            
            ws.cell(row=row, column=1, value=seat.paid_at.strftime("%Y-%m-%d %H:%M") if seat.paid_at else "")
            ws.cell(row=row, column=2, value=canal)
            ws.cell(row=row, column=3, value=producto)
            ws.cell(row=row, column=4, value=1)
            ws.cell(row=row, column=5, value=seat.total_amount)
            ws.cell(row=row, column=6, value=seat.amount_paid)
            ws.cell(row=row, column=7, value=seat.payment_status or "N/A")
            ws.cell(row=row, column=8, value="Confirmada")
            ws.cell(row=row, column=9, value=f"ASIENTO-{seat.seat_number}")
            ws.cell(row=row, column=10, value="Sistema" if canal == "Virtual" else "WhatsApp")
            row += 1
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 18
    
    def _create_reservas_detalle_sheet(self):
        """Pestaña Reservas_detalle"""
        ws = self.wb.create_sheet("Reservas_detalle")
        
        # Headers
        headers = [
            "Fecha", "Producto/Evento", "Cantidad", "Estado",
            "Vencimiento", "Contacto", "Fuente/Canal", "Observaciones"
        ]
        
        header_fill = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Datos
        row = 2
        reservas = [s for s in self.seats if s.status == SeatStatus.RESERVADO]
        
        for seat in reservas:
            producto = "Congreso Presencial"
            estado = "Confirmada" if seat.amount_paid >= seat.total_amount * 0.5 else "Pendiente"
            contacto = f"{seat.buyer_name or 'N/A'} - {seat.buyer_email or 'N/A'} - {seat.buyer_phone or 'N/A'}"
            fuente = seat.venta_origen.value if seat.venta_origen else "N/A"
            
            ws.cell(row=row, column=1, value=seat.reserved_at.strftime("%Y-%m-%d %H:%M") if seat.reserved_at else "")
            ws.cell(row=row, column=2, value=producto)
            ws.cell(row=row, column=3, value=1)
            ws.cell(row=row, column=4, value=estado)
            ws.cell(row=row, column=5, value=seat.expires_at.strftime("%Y-%m-%d %H:%M") if seat.expires_at else "N/A")
            ws.cell(row=row, column=6, value=contacto)
            ws.cell(row=row, column=7, value=fuente)
            ws.cell(row=row, column=8, value=seat.notes or "")
            row += 1
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _create_problemas_detalle_sheet(self):
        """Pestaña Problemas_detalle"""
        ws = self.wb.create_sheet("Problemas_detalle")
        
        # Headers
        headers = [
            "Fecha", "Tipo", "Severidad", "Descripción",
            "Estado", "Referencia/ID Asociado"
        ]
        
        header_fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Detectar problemas
        problemas = []
        now = datetime.now()
        
        # Reservas vencidas
        for seat in self.seats:
            if seat.status == SeatStatus.RESERVADO and seat.expires_at and seat.expires_at < now:
                problemas.append({
                    "fecha": seat.expires_at.strftime("%Y-%m-%d %H:%M"),
                    "tipo": "Reserva Vencida",
                    "severidad": "Alta",
                    "descripcion": f"Reserva del asiento {seat.seat_number} venció sin pago completo",
                    "estado": "Abierto",
                    "referencia": f"ASIENTO-{seat.seat_number}"
                })
        
        # Pagos fallidos (reservas con pago parcial muy bajo)
        for seat in self.seats:
            if seat.status == SeatStatus.RESERVADO:
                porcentaje_pagado = (seat.amount_paid / seat.total_amount * 100) if seat.total_amount > 0 else 0
                if porcentaje_pagado > 0 and porcentaje_pagado < 10:
                    problemas.append({
                        "fecha": seat.reserved_at.strftime("%Y-%m-%d %H:%M") if seat.reserved_at else "",
                        "tipo": "Pago Fallido/Incompleto",
                        "severidad": "Media",
                        "descripcion": f"Pago parcial muy bajo ({porcentaje_pagado:.1f}%) en asiento {seat.seat_number}",
                        "estado": "Abierto",
                        "referencia": f"ASIENTO-{seat.seat_number}"
                    })
        
        # Cupo inconsistente (más vendidos que capacidad)
        capacidad = self.config.get("capacity", 100)
        vendidos = len([s for s in self.seats if s.status == SeatStatus.VENDIDO])
        if vendidos > capacidad:
            problemas.append({
                "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
                "tipo": "Cupo Inconsistente",
                "severidad": "Alta",
                "descripcion": f"Vendidos ({vendidos}) exceden capacidad ({capacidad})",
                "estado": "Abierto",
                "referencia": "SISTEMA"
            })
        
        # Escribir problemas
        row = 2
        for problema in problemas:
            ws.cell(row=row, column=1, value=problema["fecha"])
            ws.cell(row=row, column=2, value=problema["tipo"])
            ws.cell(row=row, column=3, value=problema["severidad"])
            ws.cell(row=row, column=4, value=problema["descripcion"])
            ws.cell(row=row, column=5, value=problema["estado"])
            ws.cell(row=row, column=6, value=problema["referencia"])
            row += 1
        
        if row == 2:
            ws.cell(row=2, column=1, value="No se detectaron problemas")
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 25
    
    def _create_proyeccion_sheet(self):
        """Pestaña Proyeccion: Forecast de ventas"""
        ws = self.wb.create_sheet("Proyeccion")
        
        # Headers
        headers = [
            "Fecha", "Escenario Conservador", "Escenario Medio", 
            "Escenario Optimista", "Físico", "Virtual"
        ]
        
        header_fill = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
        
        # Calcular proyección
        proyeccion_dias = self.config.get("proyeccion", {}).get("dias", 28)
        fecha_fin_evento = datetime.fromisoformat(self.config.get("event", {}).get("date", "2025-12-19T08:00:00"))
        
        # Usar fecha fin de evento si es menor que proyección_dias
        dias_hasta_evento = (fecha_fin_evento - datetime.now()).days
        proyeccion_dias = min(proyeccion_dias, dias_hasta_evento) if dias_hasta_evento > 0 else proyeccion_dias
        
        # Datos históricos para forecast
        ventas_por_dia = self.metrics["ventas_por_dia"]
        ventas_fisico_por_dia = self.metrics["ventas_fisico_por_dia"]
        ventas_virtual_por_dia = self.metrics["ventas_virtual_por_dia"]
        
        # Calcular promedio móvil 7 días
        if len(ventas_por_dia) >= 7:
            ultimos_7_dias = list(ventas_por_dia.values())[-7:]
            promedio_7d = sum(ultimos_7_dias) / 7
        else:
            promedio_7d = sum(ventas_por_dia.values()) / len(ventas_por_dia) if ventas_por_dia else 0
        
        # Calcular tendencia (regresión lineal simple)
        if len(ventas_por_dia) >= 2:
            fechas = [datetime.strptime(f, "%Y-%m-%d") for f in ventas_por_dia.keys()]
            valores = list(ventas_por_dia.values())
            x = np.array([(d - fechas[0]).days for d in fechas])
            y = np.array(valores)
            
            if len(x) > 1:
                slope, intercept, _, _, _ = stats.linregress(x, y)
                tendencia_diaria = slope
            else:
                tendencia_diaria = 0
        else:
            tendencia_diaria = 0
        
        # Ratio físico/virtual
        total_fisico = sum(ventas_fisico_por_dia.values())
        total_virtual = sum(ventas_virtual_por_dia.values())
        total_ventas = total_fisico + total_virtual
        ratio_fisico = total_fisico / total_ventas if total_ventas > 0 else 0.5
        ratio_virtual = total_virtual / total_ventas if total_ventas > 0 else 0.5
        
        # Tasa de conversión reservas->venta
        tasa_conversion = self.metrics["tasa_conversion"] / 100
        
        # Generar proyección
        row = 2
        hoy = datetime.now().date()
        
        for dia in range(1, proyeccion_dias + 1):
            fecha_proy = hoy + timedelta(days=dia)
            
            # Ajustar por tendencia
            base_ventas = promedio_7d + (tendencia_diaria * dia)
            
            # Escenarios
            conservador = max(0, int(base_ventas * 0.7))
            medio = max(0, int(base_ventas))
            optimista = max(0, int(base_ventas * 1.3))
            
            # Aplicar probabilidad de conversión
            reservas_pendientes = self.metrics["reservas_total"]
            ventas_esperadas_reservas = int(reservas_pendientes * tasa_conversion / proyeccion_dias)
            
            conservador += ventas_esperadas_reservas
            medio += int(ventas_esperadas_reservas * 1.2)
            optimista += int(ventas_esperadas_reservas * 1.5)
            
            # Distribuir entre físico y virtual
            fisico = int(medio * ratio_fisico)
            virtual = int(medio * ratio_virtual)
            
            ws.cell(row=row, column=1, value=fecha_proy.strftime("%Y-%m-%d"))
            ws.cell(row=row, column=2, value=conservador)
            ws.cell(row=row, column=3, value=medio)
            ws.cell(row=row, column=4, value=optimista)
            ws.cell(row=row, column=5, value=fisico)
            ws.cell(row=row, column=6, value=virtual)
            row += 1
        
        # Ajustar columnas
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
    
    def _save_metadata(self, output_dir: Path, fecha: datetime):
        """Guardar metadata JSON con KPIs"""
        metadata = {
            "fecha_generacion": fecha.isoformat(),
            "kpis": {
                "ventas_total": self.metrics["ventas_total"],
                "ventas_fisico": self.metrics["ventas_fisico"],
                "ventas_virtual": self.metrics["ventas_virtual"],
                "reservas_total": self.metrics["reservas_total"],
                "reservas_confirmadas": self.metrics["reservas_confirmadas"],
                "tasa_conversion": self.metrics["tasa_conversion"],
                "ingresos": self.metrics["ingresos"],
                "ticket_promedio": self.metrics["ticket_promedio"],
            },
            "total_registros": len(self.seats)
        }
        
        metadata_path = output_dir / f"metadata_{fecha.strftime('%Y-%m-%d')}.json"
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(metadata, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Metadata guardada en: {metadata_path}")


def load_config(config_path: str = "reporting/config.yml") -> Dict[str, Any]:
    """Cargar configuración desde YAML"""
    config_file = Path(config_path)
    if not config_file.exists():
        raise FileNotFoundError(f"Archivo de configuración no encontrado: {config_path}")
    
    with open(config_file, 'r', encoding='utf-8') as f:
        return yaml.safe_load(f)


def main():
    """Función principal"""
    config_path = sys.argv[1] if len(sys.argv) > 1 else "reporting/config.yml"
    
    try:
        config = load_config(config_path)
        generator = SalesReportGenerator(config)
        filepath = generator.generate()
        
        print(f"\n✅ Reporte generado exitosamente: {filepath}")
        return 0
        
    except Exception as e:
        logger.error(f"Error generando reporte: {e}", exc_info=True)
        return 1


if __name__ == "__main__":
    sys.exit(main())

