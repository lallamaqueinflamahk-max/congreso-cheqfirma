"""
Tests mínimos para validar estructura del reporte
"""

import pytest
from pathlib import Path
from openpyxl import load_workbook
import json


def test_report_file_exists(tmp_path, sample_config):
    """Test que el reporte se genera correctamente"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    # Modificar output_dir temporal
    sample_config["output_dir"] = str(tmp_path / "output")
    
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    assert Path(filepath).exists(), "El archivo Excel debe existir"


def test_report_has_required_sheets(tmp_path, sample_config):
    """Test que el Excel tiene todas las pestañas requeridas"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    wb = load_workbook(filepath)
    sheet_names = wb.sheetnames
    
    required_sheets = ["Resumen", "Ventas_detalle", "Reservas_detalle", "Problemas_detalle", "Proyeccion"]
    
    for sheet_name in required_sheets:
        assert sheet_name in sheet_names, f"Falta la pestaña: {sheet_name}"


def test_ventas_detalle_has_required_columns(tmp_path, sample_config):
    """Test que Ventas_detalle tiene todas las columnas requeridas"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    wb = load_workbook(filepath)
    ws = wb["Ventas_detalle"]
    
    required_columns = [
        "Fecha", "Canal", "Producto/Evento", "Cantidad",
        "Precio", "Total", "Método de Pago", "Estado",
        "Referencia/ID", "Vendedor"
    ]
    
    headers = [cell.value for cell in ws[1]]
    
    for col in required_columns:
        assert col in headers, f"Falta la columna: {col}"


def test_reservas_detalle_has_required_columns(tmp_path, sample_config):
    """Test que Reservas_detalle tiene todas las columnas requeridas"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    wb = load_workbook(filepath)
    ws = wb["Reservas_detalle"]
    
    required_columns = [
        "Fecha", "Producto/Evento", "Cantidad", "Estado",
        "Vencimiento", "Contacto", "Fuente/Canal", "Observaciones"
    ]
    
    headers = [cell.value for cell in ws[1]]
    
    for col in required_columns:
        assert col in headers, f"Falta la columna: {col}"


def test_problemas_detalle_has_required_columns(tmp_path, sample_config):
    """Test que Problemas_detalle tiene todas las columnas requeridas"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    wb = load_workbook(filepath)
    ws = wb["Problemas_detalle"]
    
    required_columns = [
        "Fecha", "Tipo", "Severidad", "Descripción",
        "Estado", "Referencia/ID Asociado"
    ]
    
    headers = [cell.value for cell in ws[1]]
    
    for col in required_columns:
        assert col in headers, f"Falta la columna: {col}"


def test_no_critical_nulls_in_ventas(tmp_path, sample_config):
    """Test que no hay nulos críticos en Ventas_detalle"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    filepath = generator.generate()
    
    wb = load_workbook(filepath)
    ws = wb["Ventas_detalle"]
    
    # Verificar que las columnas críticas no estén vacías
    critical_columns = {
        "Fecha": 1,
        "Canal": 2,
        "Total": 6,
    }
    
    for row_idx in range(2, ws.max_row + 1):
        for col_name, col_num in critical_columns.items():
            cell_value = ws.cell(row=row_idx, column=col_num).value
            assert cell_value is not None, f"Fila {row_idx}, columna {col_name} no puede estar vacía"


def test_metadata_file_exists(tmp_path, sample_config):
    """Test que se genera el archivo metadata.json"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    generator.generate()
    
    # Buscar archivo metadata
    output_dir = Path(sample_config["output_dir"])
    metadata_files = list(output_dir.glob("metadata_*.json"))
    
    assert len(metadata_files) > 0, "Debe existir al menos un archivo metadata.json"


def test_metadata_has_required_kpis(tmp_path, sample_config):
    """Test que metadata.json tiene todos los KPIs requeridos"""
    from reporting.generate_sales_report import SalesReportGenerator
    
    sample_config["output_dir"] = str(tmp_path / "output")
    generator = SalesReportGenerator(sample_config)
    generator.generate()
    
    output_dir = Path(sample_config["output_dir"])
    metadata_files = list(output_dir.glob("metadata_*.json"))
    
    assert len(metadata_files) > 0
    
    with open(metadata_files[0], 'r', encoding='utf-8') as f:
        metadata = json.load(f)
    
    required_kpis = [
        "ventas_total", "ventas_fisico", "ventas_virtual",
        "reservas_total", "ingresos", "ticket_promedio"
    ]
    
    assert "kpis" in metadata
    for kpi in required_kpis:
        assert kpi in metadata["kpis"], f"Falta el KPI: {kpi}"


@pytest.fixture
def sample_config():
    """Configuración de prueba"""
    return {
        "event": {
            "date": "2025-12-19T08:00:00",
            "name": "Congreso ADN Humano"
        },
        "capacity": 100,
        "datasource": {
            "type": "csv",
            "csv_path": "data/seats.csv.example",
            "column_mapping": {
                "seat_id": "seat_id",
                "seat_number": "seat_number",
                "status": "status",
                "buyer_name": "buyer_name",
                "buyer_email": "buyer_email",
                "venta_origen": "venta_origen",
                "amount_paid": "amount_paid",
                "deposit_amount": "deposit_amount",
                "total_amount": "total_amount",
                "payment_status": "payment_status",
                "reserved_at": "reserved_at",
                "expires_at": "expires_at",
                "paid_at": "paid_at",
                "created_at": "created_at",
                "updated_at": "updated_at",
                "notes": "notes"
            }
        },
        "output_dir": "test_output",
        "proyeccion": {
            "dias": 28
        }
    }

